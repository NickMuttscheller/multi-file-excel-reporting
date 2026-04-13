from __future__ import annotations

import logging
import sys
import traceback
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"
LOG_DIR = BASE_DIR / "log"

LOG_FILE = LOG_DIR / "reporting_run.log"
OUTPUT_WORKBOOK = OUTPUT_DIR / "sales_reporting_pack.xlsx"
OUTPUT_CLEAN_CSV = OUTPUT_DIR / "cleaned_master_data.csv"
OUTPUT_QUALITY_CSV = OUTPUT_DIR / "data_quality_issues.csv"
OUTPUT_SUMMARY_TXT = OUTPUT_DIR / "processing_summary.txt"

SUPPORTED_EXTENSIONS = {".csv", ".xlsx", ".xls"}

CANONICAL_COLUMNS = [
    "date",
    "region",
    "salesperson",
    "customer_name",
    "product",
    "category",
    "units_sold",
    "unit_price",
    "revenue",
    "status",
]

REQUIRED_COLUMNS = [
    "date",
    "region",
    "salesperson",
    "customer_name",
    "product",
    "category",
    "units_sold",
    "unit_price",
]

COLUMN_ALIASES = {
    "order_date": "date",
    "transaction_date": "date",
    "sale_date": "date",
    "sales_date": "date",
    "territory": "region",
    "sales_region": "region",
    "rep": "salesperson",
    "sales_rep": "salesperson",
    "account_manager": "salesperson",
    "customer": "customer_name",
    "client": "customer_name",
    "client_name": "customer_name",
    "item": "product",
    "product_name": "product",
    "product_category": "category",
    "qty": "units_sold",
    "quantity": "units_sold",
    "units": "units_sold",
    "price": "unit_price",
    "price_per_unit": "unit_price",
    "sales_amount": "revenue",
    "amount": "revenue",
    "order_status": "status",
    "deal_status": "status",
}


@dataclass
class FileProcessingRecord:
    file_name: str
    file_type: str
    status: str
    rows_read: int = 0
    rows_after_standardization: int = 0
    rows_after_cleaning: int = 0
    fully_empty_rows_removed: int = 0
    duplicates_removed: int = 0
    invalid_dates_removed: int = 0
    invalid_numeric_rows_removed: int = 0
    missing_required_rows_removed: int = 0
    notes: str = ""


@dataclass
class PipelineMetrics:
    total_files_found: int = 0
    total_files_processed: int = 0
    total_files_skipped: int = 0
    raw_combined_row_count: int = 0
    fully_empty_rows_removed: int = 0
    duplicates_removed: int = 0
    invalid_dates_removed: int = 0
    invalid_numeric_rows_removed: int = 0
    missing_required_rows_removed: int = 0
    final_cleaned_row_count: int = 0
    files_processed: list[str] = field(default_factory=list)
    files_skipped: list[str] = field(default_factory=list)


def setup_directories() -> None:
    """Create required project directories."""
    for directory in (INPUT_DIR, OUTPUT_DIR, LOG_DIR):
        directory.mkdir(parents=True, exist_ok=True)


def setup_logging() -> None:
    """Configure production-style file and console logging."""
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    formatter = logging.Formatter(
        "%(asctime)s | %(levelname)-8s | %(name)s | %(message)s"
    )

    file_handler = logging.FileHandler(LOG_FILE, encoding="utf-8")
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(formatter)

    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)


def normalize_column_name(column_name: Any) -> str:
    """Normalize raw column names into internal snake_case keys."""
    value = str(column_name).strip().lower()
    value = value.replace("%", "percent")
    value = value.replace("/", "_")
    value = value.replace("-", "_")
    value = value.replace(" ", "_")
    while "__" in value:
        value = value.replace("__", "_")
    return COLUMN_ALIASES.get(value, value)


def read_input_file(file_path: Path) -> pd.DataFrame:
    """Read a supported input file into a DataFrame."""
    suffix = file_path.suffix.lower()

    if suffix == ".csv":
        return pd.read_csv(file_path)

    if suffix in {".xlsx", ".xls"}:
        return pd.read_excel(file_path)

    raise ValueError(f"Unsupported file type: {suffix}")


def discover_input_files() -> list[Path]:
    """Return supported files from the input directory."""
    all_files = [path for path in INPUT_DIR.iterdir() if path.is_file()]
    supported = [path for path in all_files if path.suffix.lower() in SUPPORTED_EXTENSIONS]
    unsupported = [path for path in all_files if path.suffix.lower() not in SUPPORTED_EXTENSIONS]

    for file_path in supported:
        logging.info("Discovered supported input file: %s", file_path.name)

    for file_path in unsupported:
        logging.warning("Skipping unsupported file: %s", file_path.name)

    return sorted(supported)


def apply_text_normalization(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    """Trim and standardize textual fields without damaging missing values."""
    title_case_columns = {"region", "salesperson", "customer_name", "product", "category"}
    upper_case_status_values = {"won", "open", "closed"}
    standardized_status_map = {
        "completed": "Completed",
        "complete": "Completed",
        "closed": "Completed",
        "paid": "Completed",
        "pending": "Pending",
        "open": "Pending",
        "cancelled": "Cancelled",
        "canceled": "Cancelled",
        "refunded": "Refunded",
    }

    for column in columns:
        if column not in df.columns:
            continue

        series = df[column].astype("string").str.strip()

        series = series.replace(
            {
                "": pd.NA,
                "nan": pd.NA,
                "none": pd.NA,
                "null": pd.NA,
                "n/a": pd.NA,
                "na": pd.NA,
            }
        )

        if column in title_case_columns:
            series = series.str.replace(r"\s+", " ", regex=True).str.title()

        if column == "status":
            lowered = series.str.lower()
            series = lowered.map(standardized_status_map).fillna(series.str.title())

        if column == "region":
            series = series.str.replace(r"\s+", " ", regex=True).str.title()

        if column == "status":
            mask = series.astype("string").str.lower().isin(upper_case_status_values)
            series = series.mask(mask, series.str.title())

        df[column] = series

    return df


def standardize_dataframe(raw_df: pd.DataFrame, source_file: str) -> pd.DataFrame:
    """Standardize schema and append metadata columns."""
    df = raw_df.copy()
    df.columns = [normalize_column_name(column) for column in df.columns]

    duplicate_columns = df.columns[df.columns.duplicated()].tolist()
    if duplicate_columns:
        raise ValueError(
            f"Duplicate columns found after normalization in {source_file}: {duplicate_columns}"
        )

    for column in CANONICAL_COLUMNS:
        if column not in df.columns:
            df[column] = pd.NA

    df = df[CANONICAL_COLUMNS].copy()
    df["source_file"] = source_file
    return df


def clean_single_dataframe(
    df: pd.DataFrame, file_name: str
) -> tuple[pd.DataFrame, FileProcessingRecord, pd.DataFrame]:
    """Clean a standardized single-file DataFrame and return valid rows, metrics, and rejected rows."""
    record = FileProcessingRecord(
        file_name=file_name,
        file_type=Path(file_name).suffix.lower(),
        status="Processed",
        rows_read=len(df),
    )

    quality_issue_frames: list[pd.DataFrame] = []

    starting_rows = len(df)
    df = df.dropna(how="all").copy()
    record.fully_empty_rows_removed = starting_rows - len(df)
    record.rows_after_standardization = len(df)

    df = apply_text_normalization(
        df,
        ["region", "salesperson", "customer_name", "product", "category", "status"],
    )

    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    invalid_date_mask = df["date"].isna()
    record.invalid_dates_removed = int(invalid_date_mask.sum())

    if invalid_date_mask.any():
        rejected = df.loc[invalid_date_mask].copy()
        rejected["rejection_reason"] = "Invalid date"
        quality_issue_frames.append(rejected)

    df["units_sold"] = pd.to_numeric(df["units_sold"], errors="coerce")
    df["unit_price"] = pd.to_numeric(df["unit_price"], errors="coerce")
    df["revenue"] = pd.to_numeric(df["revenue"], errors="coerce")

    numeric_invalid_mask = (
        df["units_sold"].isna()
        | df["unit_price"].isna()
        | (df["units_sold"] < 0)
        | (df["unit_price"] < 0)
    )
    record.invalid_numeric_rows_removed = int(numeric_invalid_mask.sum())

    if numeric_invalid_mask.any():
        rejected = df.loc[numeric_invalid_mask].copy()
        rejected["rejection_reason"] = "Invalid numeric values"
        quality_issue_frames.append(rejected)

    valid_numeric_mask = ~numeric_invalid_mask
    df.loc[valid_numeric_mask, "revenue"] = (
        df.loc[valid_numeric_mask, "units_sold"] * df.loc[valid_numeric_mask, "unit_price"]
    ).round(2)

    required_mask = pd.Series(False, index=df.index)
    for column in ["region", "salesperson", "customer_name", "product", "category"]:
        required_mask = required_mask | df[column].isna()

    required_mask = required_mask | df["date"].isna()
    required_mask = required_mask | df["units_sold"].isna()
    required_mask = required_mask | df["unit_price"].isna()

    record.missing_required_rows_removed = int(required_mask.sum())

    if required_mask.any():
        rejected = df.loc[required_mask].copy()
        rejected["rejection_reason"] = "Missing required fields"
        quality_issue_frames.append(rejected)

    valid_mask = ~(invalid_date_mask | numeric_invalid_mask | required_mask)
    cleaned = df.loc[valid_mask].copy()

    pre_dedup_rows = len(cleaned)
    dedupe_subset = [
        "date",
        "region",
        "salesperson",
        "customer_name",
        "product",
        "category",
        "units_sold",
        "unit_price",
        "revenue",
        "status",
    ]
    cleaned = cleaned.drop_duplicates(subset=dedupe_subset).copy()
    record.duplicates_removed = pre_dedup_rows - len(cleaned)
    record.rows_after_cleaning = len(cleaned)

    cleaned["date"] = pd.to_datetime(cleaned["date"]).dt.normalize()
    cleaned["month"] = cleaned["date"].dt.to_period("M").astype(str)

    rejected_rows = (
        pd.concat(quality_issue_frames, ignore_index=True)
        if quality_issue_frames
        else pd.DataFrame(columns=list(df.columns) + ["rejection_reason"])
    )

    note_parts = []
    if record.fully_empty_rows_removed:
        note_parts.append(f"empty rows removed: {record.fully_empty_rows_removed}")
    if record.invalid_dates_removed:
        note_parts.append(f"invalid dates removed: {record.invalid_dates_removed}")
    if record.invalid_numeric_rows_removed:
        note_parts.append(f"invalid numeric rows removed: {record.invalid_numeric_rows_removed}")
    if record.missing_required_rows_removed:
        note_parts.append(f"rows missing required fields: {record.missing_required_rows_removed}")
    if record.duplicates_removed:
        note_parts.append(f"duplicates removed: {record.duplicates_removed}")
    record.notes = "; ".join(note_parts) if note_parts else "No issues detected"

    logging.info(
        "Processed %s | read=%s | empty_removed=%s | invalid_dates=%s | invalid_numeric=%s | "
        "missing_required=%s | duplicates_removed=%s | final_rows=%s",
        file_name,
        record.rows_read,
        record.fully_empty_rows_removed,
        record.invalid_dates_removed,
        record.invalid_numeric_rows_removed,
        record.missing_required_rows_removed,
        record.duplicates_removed,
        record.rows_after_cleaning,
    )

    return cleaned, record, rejected_rows


def process_all_input_files() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, PipelineMetrics]:
    """Read, clean, and combine all supported input files."""
    files = discover_input_files()
    metrics = PipelineMetrics(total_files_found=len(files))

    processed_frames: list[pd.DataFrame] = []
    rejected_frames: list[pd.DataFrame] = []
    file_summary_records: list[dict[str, Any]] = []

    if not files:
        raise FileNotFoundError(
            f"No supported input files were found in: {INPUT_DIR}"
        )

    for file_path in files:
        try:
            logging.info("Reading file: %s", file_path.name)
            raw_df = read_input_file(file_path)
            standardized_df = standardize_dataframe(raw_df, file_path.name)
            cleaned_df, record, rejected_rows = clean_single_dataframe(
                standardized_df,
                file_path.name,
            )

            processed_frames.append(cleaned_df)

            if not rejected_rows.empty:
                rejected_frames.append(rejected_rows)

            file_summary_records.append(record.__dict__)
            metrics.total_files_processed += 1
            metrics.files_processed.append(file_path.name)

        except Exception as exc:
            logging.exception("Failed to process file: %s", file_path.name)
            metrics.total_files_skipped += 1
            metrics.files_skipped.append(file_path.name)

            error_record = FileProcessingRecord(
                file_name=file_path.name,
                file_type=file_path.suffix.lower(),
                status="Skipped",
                notes=str(exc),
            )
            file_summary_records.append(error_record.__dict__)

    file_summary_df = pd.DataFrame(file_summary_records)

    if processed_frames:
        combined_df = pd.concat(processed_frames, ignore_index=True)
    else:
        combined_df = pd.DataFrame(columns=CANONICAL_COLUMNS + ["source_file", "month"])

    if rejected_frames:
        rejected_df = pd.concat(rejected_frames, ignore_index=True)
    else:
        rejected_df = pd.DataFrame()

    metrics.raw_combined_row_count = int(file_summary_df["rows_read"].fillna(0).sum())
    metrics.fully_empty_rows_removed = int(file_summary_df["fully_empty_rows_removed"].fillna(0).sum())
    metrics.duplicates_removed = int(file_summary_df["duplicates_removed"].fillna(0).sum())
    metrics.invalid_dates_removed = int(file_summary_df["invalid_dates_removed"].fillna(0).sum())
    metrics.invalid_numeric_rows_removed = int(
        file_summary_df["invalid_numeric_rows_removed"].fillna(0).sum()
    )
    metrics.missing_required_rows_removed = int(
        file_summary_df["missing_required_rows_removed"].fillna(0).sum()
    )
    metrics.final_cleaned_row_count = len(combined_df)

    if combined_df.empty:
        raise ValueError("No valid records remained after processing all files.")

    return combined_df, rejected_df, file_summary_df, metrics


def build_kpi_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Build a compact KPI summary table."""
    total_revenue = float(df["revenue"].sum())
    total_units = int(df["units_sold"].sum())
    total_transactions = int(len(df))
    avg_transaction = total_revenue / total_transactions if total_transactions else 0.0

    revenue_by_salesperson = (
        df.groupby("salesperson", dropna=False)["revenue"].sum().sort_values(ascending=False)
    )
    revenue_by_region = df.groupby("region", dropna=False)["revenue"].sum().sort_values(ascending=False)
    revenue_by_product = df.groupby("product", dropna=False)["revenue"].sum().sort_values(ascending=False)

    summary_rows = [
        ("Total Revenue", total_revenue),
        ("Total Units Sold", total_units),
        ("Total Transactions", total_transactions),
        ("Average Revenue Per Transaction", round(avg_transaction, 2)),
        ("Top Salesperson", revenue_by_salesperson.index[0] if not revenue_by_salesperson.empty else "N/A"),
        ("Top Salesperson Revenue", float(revenue_by_salesperson.iloc[0]) if not revenue_by_salesperson.empty else 0.0),
        ("Top Region", revenue_by_region.index[0] if not revenue_by_region.empty else "N/A"),
        ("Top Region Revenue", float(revenue_by_region.iloc[0]) if not revenue_by_region.empty else 0.0),
        ("Top Product", revenue_by_product.index[0] if not revenue_by_product.empty else "N/A"),
        ("Top Product Revenue", float(revenue_by_product.iloc[0]) if not revenue_by_product.empty else 0.0),
    ]

    return pd.DataFrame(summary_rows, columns=["KPI", "Value"])


def build_aggregation_tables(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """Create business aggregation tables used in the report and charts."""
    aggregation_tables = {
        "Revenue by Region": (
            df.groupby("region", as_index=False)["revenue"]
            .sum()
            .sort_values("revenue", ascending=False)
            .reset_index(drop=True)
        ),
        "Revenue by Product": (
            df.groupby("product", as_index=False)["revenue"]
            .sum()
            .sort_values("revenue", ascending=False)
            .reset_index(drop=True)
        ),
        "Revenue by Category": (
            df.groupby("category", as_index=False)["revenue"]
            .sum()
            .sort_values("revenue", ascending=False)
            .reset_index(drop=True)
        ),
        "Revenue by Salesperson": (
            df.groupby("salesperson", as_index=False)["revenue"]
            .sum()
            .sort_values("revenue", ascending=False)
            .reset_index(drop=True)
        ),
        "Monthly Revenue Trend": (
            df.groupby("month", as_index=False)["revenue"]
            .sum()
            .sort_values("month")
            .reset_index(drop=True)
        ),
    }
    return aggregation_tables


def build_processing_summary_df(metrics: PipelineMetrics) -> pd.DataFrame:
    """Create a processing summary table."""
    rows = [
        ("Total Files Found", metrics.total_files_found),
        ("Files Successfully Processed", metrics.total_files_processed),
        ("Files Skipped", metrics.total_files_skipped),
        ("Raw Combined Row Count", metrics.raw_combined_row_count),
        ("Fully Empty Rows Removed", metrics.fully_empty_rows_removed),
        ("Duplicates Removed", metrics.duplicates_removed),
        ("Invalid Dates Removed", metrics.invalid_dates_removed),
        ("Invalid Numeric Rows Removed", metrics.invalid_numeric_rows_removed),
        ("Rows Missing Required Fields", metrics.missing_required_rows_removed),
        ("Final Cleaned Row Count", metrics.final_cleaned_row_count),
        ("Processed Files", ", ".join(metrics.files_processed) if metrics.files_processed else "None"),
        ("Skipped Files", ", ".join(metrics.files_skipped) if metrics.files_skipped else "None"),
    ]
    return pd.DataFrame(rows, columns=["Metric", "Value"])


def write_excel_report(
    clean_df: pd.DataFrame,
    kpi_df: pd.DataFrame,
    aggregation_tables: dict[str, pd.DataFrame],
    file_summary_df: pd.DataFrame,
    processing_summary_df: pd.DataFrame,
) -> None:
    """Write the polished Excel reporting workbook."""
    with pd.ExcelWriter(OUTPUT_WORKBOOK, engine="openpyxl") as writer:
        clean_df.to_excel(writer, sheet_name="Cleaned Master Data", index=False)
        kpi_df.to_excel(writer, sheet_name="KPI Summary", index=False, startrow=1)
        processing_summary_df.to_excel(writer, sheet_name="KPI Summary", index=False, startrow=14)

        start_row = 1
        for table_name, table_df in aggregation_tables.items():
            table_df.to_excel(
                writer,
                sheet_name="Aggregations",
                index=False,
                startrow=start_row,
            )
            start_row += len(table_df) + 4

        file_summary_df.to_excel(writer, sheet_name="File Processing Summary", index=False)
        chart_anchor_df = aggregation_tables["Revenue by Region"]
        chart_anchor_df.to_excel(writer, sheet_name="Charts", index=False, startrow=1)

        aggregation_tables["Revenue by Product"].to_excel(
            writer, sheet_name="Charts", index=False, startrow=1, startcol=4
        )
        aggregation_tables["Monthly Revenue Trend"].to_excel(
            writer, sheet_name="Charts", index=False, startrow=20, startcol=0
        )

    wb = load_workbook(OUTPUT_WORKBOOK)
    style_workbook(wb)
    add_charts_to_workbook(wb, aggregation_tables)
    wb.save(OUTPUT_WORKBOOK)


def style_workbook(workbook) -> None:
    """Apply consistent report styling across worksheets."""
    header_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(color="FFFFFF", bold=True)
    section_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    currency_columns = {"H", "I"}
    integer_columns = {"G"}

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]

        if sheet_name == "KPI Summary":
            ws["A1"] = "KPI Summary"
            ws["A14"] = "Processing Summary"
            ws["A1"].font = Font(bold=True, size=13)
            ws["A14"].font = Font(bold=True, size=13)

        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        for row_idx in range(1, ws.max_row + 1):
            values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
            if any(value is not None for value in values):
                if row_idx == 1 or sheet_name in {"Cleaned Master Data", "File Processing Summary"} and row_idx == 1:
                    for cell in ws[row_idx]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                elif sheet_name == "Aggregations" and ws.cell(row=row_idx, column=1).value in {
                    "region",
                    "product",
                    "category",
                    "salesperson",
                    "month",
                }:
                    for cell in ws[row_idx]:
                        cell.fill = section_fill
                        cell.font = section_font

        if ws.max_row >= 1:
            ws.freeze_panes = "A2"

        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)

            for cell in ws[column_letter]:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))

            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 35)

        if sheet_name == "Cleaned Master Data":
            for cell in ws["A"]:
                if cell.row > 1:
                    cell.number_format = "yyyy-mm-dd"

            for col in currency_columns:
                for cell in ws[col]:
                    if cell.row > 1:
                        cell.number_format = '$#,##0.00'

            for col in integer_columns:
                for cell in ws[col]:
                    if cell.row > 1:
                        cell.number_format = "#,##0"

        if sheet_name in {"Aggregations", "Charts"}:
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        if ws.cell(row=1, column=cell.column).value == "revenue":
                            cell.number_format = '$#,##0.00'

        if sheet_name == "KPI Summary":
            for row in range(2, 12):
                value_cell = ws.cell(row=row, column=2)
                if isinstance(value_cell.value, (int, float)):
                    if row in {2, 5, 8, 10}:
                        value_cell.number_format = '$#,##0.00'
                    elif row == 3 or row == 4:
                        value_cell.number_format = "#,##0"

        ws.sheet_view.showGridLines = False


def add_charts_to_workbook(workbook, aggregation_tables: dict[str, pd.DataFrame]) -> None:
    """Create openpyxl charts on the Charts worksheet."""
    ws = workbook["Charts"]

    region_rows = len(aggregation_tables["Revenue by Region"]) + 1
    product_rows = len(aggregation_tables["Revenue by Product"]) + 1
    monthly_rows = len(aggregation_tables["Monthly Revenue Trend"]) + 20

    region_chart = BarChart()
    region_chart.title = "Revenue by Region"
    region_chart.y_axis.title = "Revenue"
    region_chart.x_axis.title = "Region"
    region_chart.height = 8
    region_chart.width = 14

    region_data = Reference(ws, min_col=2, min_row=1, max_row=region_rows)
    region_categories = Reference(ws, min_col=1, min_row=2, max_row=region_rows)
    region_chart.add_data(region_data, titles_from_data=True)
    region_chart.set_categories(region_categories)
    ws.add_chart(region_chart, "J2")

    product_chart = BarChart()
    product_chart.title = "Revenue by Product"
    product_chart.y_axis.title = "Revenue"
    product_chart.x_axis.title = "Product"
    product_chart.height = 8
    product_chart.width = 16

    product_data = Reference(ws, min_col=6, min_row=1, max_row=product_rows)
    product_categories = Reference(ws, min_col=5, min_row=2, max_row=product_rows)
    product_chart.add_data(product_data, titles_from_data=True)
    product_chart.set_categories(product_categories)
    ws.add_chart(product_chart, "J20")

    trend_chart = LineChart()
    trend_chart.title = "Monthly Revenue Trend"
    trend_chart.y_axis.title = "Revenue"
    trend_chart.x_axis.title = "Month"
    trend_chart.height = 8
    trend_chart.width = 16
    trend_chart.style = 2

    trend_data = Reference(ws, min_col=2, min_row=20, max_row=monthly_rows)
    trend_categories = Reference(ws, min_col=1, min_row=21, max_row=monthly_rows)
    trend_chart.add_data(trend_data, titles_from_data=True)
    trend_chart.set_categories(trend_categories)
    ws.add_chart(trend_chart, "J38")


def write_supporting_outputs(
    clean_df: pd.DataFrame,
    rejected_df: pd.DataFrame,
    file_summary_df: pd.DataFrame,
    processing_summary_df: pd.DataFrame,
    kpi_df: pd.DataFrame,
) -> None:
    """Write supporting CSV and text outputs."""
    clean_df.to_csv(OUTPUT_CLEAN_CSV, index=False)

    if not rejected_df.empty:
        rejected_df.to_csv(OUTPUT_QUALITY_CSV, index=False)

    lines: list[str] = []
    lines.append("MULTI-FILE EXCEL REPORTING PIPELINE")
    lines.append("=" * 40)
    lines.append("")
    lines.append("Processing Summary")
    lines.append("-" * 40)
    for _, row in processing_summary_df.iterrows():
        lines.append(f"{row['Metric']}: {row['Value']}")

    lines.append("")
    lines.append("KPI Summary")
    lines.append("-" * 40)
    for _, row in kpi_df.iterrows():
        lines.append(f"{row['KPI']}: {row['Value']}")

    lines.append("")
    lines.append("File-Level Processing Results")
    lines.append("-" * 40)
    for _, row in file_summary_df.iterrows():
        lines.append(
            f"{row['file_name']} | status={row['status']} | rows_read={row['rows_read']} | "
            f"rows_after_cleaning={row['rows_after_cleaning']} | notes={row['notes']}"
        )

    OUTPUT_SUMMARY_TXT.write_text("\n".join(lines), encoding="utf-8")


def log_pipeline_summary(metrics: PipelineMetrics) -> None:
    """Write a concise pipeline summary to the log."""
    logging.info("Pipeline summary")
    logging.info("Total files found: %s", metrics.total_files_found)
    logging.info("Files successfully processed: %s", metrics.total_files_processed)
    logging.info("Files skipped: %s", metrics.total_files_skipped)
    logging.info("Raw combined row count: %s", metrics.raw_combined_row_count)
    logging.info("Fully empty rows removed: %s", metrics.fully_empty_rows_removed)
    logging.info("Duplicates removed: %s", metrics.duplicates_removed)
    logging.info("Invalid dates removed: %s", metrics.invalid_dates_removed)
    logging.info("Invalid numeric rows removed: %s", metrics.invalid_numeric_rows_removed)
    logging.info("Rows missing required fields: %s", metrics.missing_required_rows_removed)
    logging.info("Final cleaned row count: %s", metrics.final_cleaned_row_count)


def main() -> None:
    """Run the reporting pipeline end-to-end."""
    setup_directories()
    setup_logging()
    logging.info("Starting multi-file Excel reporting pipeline")
    logging.info("Base directory: %s", BASE_DIR)
    logging.info("Input directory: %s", INPUT_DIR)
    logging.info("Output directory: %s", OUTPUT_DIR)
    logging.info("Log file: %s", LOG_FILE)

    try:
        clean_df, rejected_df, file_summary_df, metrics = process_all_input_files()

        kpi_df = build_kpi_summary(clean_df)
        aggregation_tables = build_aggregation_tables(clean_df)
        processing_summary_df = build_processing_summary_df(metrics)

        write_excel_report(
            clean_df=clean_df,
            kpi_df=kpi_df,
            aggregation_tables=aggregation_tables,
            file_summary_df=file_summary_df,
            processing_summary_df=processing_summary_df,
        )

        write_supporting_outputs(
            clean_df=clean_df,
            rejected_df=rejected_df,
            file_summary_df=file_summary_df,
            processing_summary_df=processing_summary_df,
            kpi_df=kpi_df,
        )

        log_pipeline_summary(metrics)
        logging.info("Workbook created: %s", OUTPUT_WORKBOOK)
        logging.info("Clean CSV created: %s", OUTPUT_CLEAN_CSV)
        if not rejected_df.empty:
            logging.info("Data quality export created: %s", OUTPUT_QUALITY_CSV)
        logging.info("Summary report created: %s", OUTPUT_SUMMARY_TXT)
        logging.info("Reporting pipeline completed successfully")

    except Exception as exc:
        logging.error("Pipeline failed: %s", exc)
        logging.error(traceback.format_exc())
        raise


if __name__ == "__main__":
    main()